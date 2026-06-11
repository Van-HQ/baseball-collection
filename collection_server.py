#!/usr/bin/env python3
"""
collection_server.py — Local API server for the baseball collection dashboard.

Run this once before opening baseball_collection.html:
    python3 collection_server.py

Press Ctrl+C to stop.
"""

import json, re, sys, time, subprocess
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse, parse_qs, quote
from pathlib import Path
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

PORT = 5055
XLSX = Path.home() / 'Downloads' / 'Baseball Collection.xlsx'
HERE = Path(__file__).parent

# ── Card number prefix → SCP set slug ─────────────────────────────────────────
# Used to construct a direct card URL (most reliable approach)
PREFIX_SET = {
    'BDC':  'bowman-draft-chrome',
    'BDN':  'bowman-draft-chrome',       # Draft Night insert
    'BCP':  'bowman-chrome',
    'BPA':  'bowman-chrome',             # Chrome Prospect Auto
    'CPA':  'bowman-chrome',
    'BCPA': 'bowman-chrome',
    'BP':   'bowman',
    'PP':   'bowman-chrome',             # Prized Prospects
    'BTP':  'bowman-chrome',             # Top Prospects
    'BST':  'bowman-chrome',             # Star of Tomorrow
    'ROY':  'bowman-chrome',
    'GL':   'bowman-chrome',
    'RR':   'bowman-chrome',
    'VIP':  'bowman-chrome',
    'ES':   'bowman-chrome',
    'UR':   'bowman-chrome',
    'PC':   'bowman-chrome',
    'NT':   'topps-chrome-update',       # Night Terrors
    'LS':   'topps-cosmic-chrome',       # Lunar Surface
    'BT':   'topps-cosmic-chrome',       # Black Ties
}

def slugify(s):
    s = s.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', ' ', s)
    s = re.sub(r'\s+', '-', s.strip())
    return re.sub(r'-+', '-', s).strip('-')

def normalize(s):
    return re.sub(r'[^a-z0-9\s]', ' ', s.lower())

def parse_price(s):
    if not s: return None
    m = re.search(r'[\d,]+\.?\d*', s.replace(',', ''))
    try:    return float(m.group()) if m else None
    except: return None

def card_prefix(card_no):
    """Extract the letter prefix from a card number like BDC-166 → BDC."""
    m = re.match(r'^([A-Za-z]+)', (card_no or '').strip())
    return m.group(1).upper() if m else None


# ── Strategy 1: direct URL ─────────────────────────────────────────────────────

def direct_card_url(player, parallel, year, card_no):
    """
    Attempt to build the SCP direct card page URL.
    Pattern: /game/baseball-cards-{year}-{set-slug}/{player-slug}-{parallel-slug}-{card-no-slug}
    """
    prefix = card_prefix(card_no)
    if not prefix or prefix not in PREFIX_SET:
        return None

    set_name = PREFIX_SET[prefix]
    set_slug = f"baseball-cards-{year}-{set_name}"

    # Strip the serial number from the parallel for the slug
    # e.g. "Green Geometric /99" → "green-geometric"
    parallel_clean = re.sub(r'/\d+', '', parallel)
    parallel_slug  = slugify(parallel_clean)

    player_slug = slugify(player)
    card_slug   = card_no.lower().replace('#', '').strip()

    # SCP slug: player + parallel keywords + card number
    card_page_slug = f"{player_slug}-{parallel_slug}-{card_slug}"

    # Also build the ?q= param SCP uses (same format as user-provided URL)
    q = quote(f"{player} {parallel} {card_no}")
    return f"https://www.sportscardspro.com/game/{set_slug}/{card_page_slug}?q={q}"


def scrape_card_page(page, url):
    """Scrape a direct SCP card page (game/ URL) for prices."""
    print(f"  → Direct URL: {url}")
    try:
        page.goto(url, timeout=25000, wait_until="domcontentloaded")
        page.wait_for_timeout(2500)
        soup = BeautifulSoup(page.content(), "html.parser")

        # SCP card pages show prices in various formats — try the main price table first
        result = _parse_card_page(soup, url)
        if result:
            return result

        # If we got a 404 or redirect, fall through to None
        title_tag = soup.find('title')
        if title_tag and ('not found' in title_tag.text.lower() or '404' in title_tag.text):
            print("  ✗ Direct URL returned 404")
            return None

        # Page loaded but prices not found in expected location
        print("  – Direct page loaded but prices not parsed")
        return None
    except Exception as e:
        print(f"  ✗ Direct URL error: {e}")
        return None


def _parse_card_page(soup, url):
    """
    Parse prices from a direct /game/ card page.
    SCP shows prices in a table with rows for RAW, PSA 9, PSA 10.
    """
    # Look for price rows — SCP typically has a pricing section
    prices = {}

    # Try to find the grade/price table
    tables = soup.find_all('table')
    for table in tables:
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            if len(cells) < 2:
                continue
            label = cells[0].get_text(strip=True).lower()
            val   = parse_price(cells[-1].get_text(strip=True))
            if val is None:
                continue
            if 'ungraded' in label or 'raw' in label:
                prices['ungraded'] = val
            elif 'psa 10' in label or 'psa10' in label:
                prices['psa10'] = val
            elif 'psa 9' in label or 'grade 9' in label or 'psa9' in label:
                prices['grade9'] = val

    # Also try definition list / price spans common on SCP card pages
    for el in soup.find_all(class_=re.compile(r'price|grade|ungraded', re.I)):
        txt = el.get_text(strip=True).lower()
        val = parse_price(txt)
        if val and 'ungraded' not in prices and ('ungraded' in txt or 'raw' in txt):
            prices['ungraded'] = val

    if prices:
        # Get card title from page
        h1 = soup.find('h1')
        title = h1.get_text(strip=True) if h1 else ''
        return {
            'title':    title,
            'set':      '',
            'ungraded': prices.get('ungraded'),
            'grade9':   prices.get('grade9'),
            'psa10':    prices.get('psa10'),
            'url':      url,
        }
    return None


# ── Strategy 2: search page ───────────────────────────────────────────────────

def scrape_search_page(page, player, parallel, card_no):
    """
    Fall back to search-products page.
    When card_no is provided, search player + card_no only (more precise).
    """
    if card_no:
        # Searching by card number is much more targeted than full query
        q = f"{player} {card_no}"
    else:
        q = f"{player} {parallel}"

    url = (
        f"https://www.sportscardspro.com/search-products"
        f"?type=prices&q={quote(q)}&sport=baseball-cards"
        f"&rookies-only=false&exclude-variants=false&show-images=true"
    )
    print(f"  → Search fallback: {q}")

    try:
        page.goto(url, timeout=25000, wait_until="domcontentloaded")
        page.wait_for_timeout(2500)
        soup  = BeautifulSoup(page.content(), "html.parser")
        table = soup.find("table")
        if not table:
            print("  ✗ No table on search page")
            return None

        rows = table.find_all("tr")
        if len(rows) < 2:
            return None

        hdrs = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]

        # ── Format A: classic search results (Title / Set / grade columns) ──
        if "Title" in hdrs and "Ungraded" in hdrs:
            ti   = hdrs.index("Title")
            si   = hdrs.index("Set") if "Set" in hdrs else None
            ui   = hdrs.index("Ungraded")
            g9i  = hdrs.index("Grade 9") if "Grade 9" in hdrs else None
            p10i = hdrs.index("PSA 10")  if "PSA 10" in hdrs else None

            player_parts   = player.lower().split()
            parallel_norm  = normalize(parallel) if parallel else ''
            parallel_words = [w for w in parallel_norm.split() if len(w) >= 2]
            cn_clean       = normalize(card_no) if card_no else None

            best, best_score = None, 9999
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if len(cells) <= ti:
                    continue
                title_raw = cells[ti].get_text(separator=" ", strip=True)
                title_n   = normalize(title_raw)
                if not all(p in title_n for p in player_parts):
                    continue
                score = 0
                if cn_clean:
                    cn_words = cn_clean.split()
                    if all(w in title_n for w in cn_words):   score -= 100
                    elif any(w in title_n for w in cn_words): score -= 30
                if parallel_words:
                    score += sum(1 for w in parallel_words if w not in title_n) * 10
                if score < best_score:
                    best_score = score
                    a_tag = cells[ti].find("a", href=True)
                    href  = a_tag["href"] if a_tag else None
                    if href and href.startswith("/"):
                        href = "https://www.sportscardspro.com" + href
                    best = {
                        "title":    title_raw,
                        "set":      cells[si].get_text(strip=True) if si is not None else '',
                        "ungraded": parse_price(cells[ui].get_text(strip=True)) if ui is not None else None,
                        "grade9":   parse_price(cells[g9i].get_text(strip=True)) if g9i is not None else None,
                        "psa10":    parse_price(cells[p10i].get_text(strip=True)) if p10i is not None else None,
                        "url":      href,
                    }
            if best:
                print(f"  ✓ Search match (score={best_score}): {best['title']}")
            else:
                print("  – No match in search results")
            return best

        # ── Format B: grade-price table (SCP card/redirect page) ──
        # Headers look like: ['Ungraded', 'Grade 7', 'Grade 8', 'Grade 9', 'Grade 9.5', 'PSA 10', ...]
        if "Ungraded" in hdrs or "PSA 10" in hdrs:
            print(f"  → Grade-table format detected: {hdrs}")
            ui   = hdrs.index("Ungraded") if "Ungraded" in hdrs else None
            g9i  = hdrs.index("Grade 9")  if "Grade 9"  in hdrs else None
            p10i = hdrs.index("PSA 10")   if "PSA 10"   in hdrs else None

            # Card title from page heading; URL is the current page
            h1        = soup.find("h1")
            page_title = h1.get_text(strip=True) if h1 else player
            current_url = page.url

            # Each data row may represent a variant — pick the one whose row label
            # (first cell, if non-numeric) best matches the requested parallel.
            parallel_norm  = normalize(parallel) if parallel else ''
            parallel_words = [w for w in parallel_norm.split() if len(w) >= 2]

            best, best_score = None, 9999
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if not cells:
                    continue

                # First cell might be a row-label (variant name) or a price
                row_label  = cells[0].get_text(strip=True)
                label_is_price = parse_price(row_label) is not None

                # Determine actual price cell indices
                price_offset = 0 if label_is_price else 1
                def cell_price(idx):
                    actual = (idx if label_is_price else idx + price_offset) if idx is not None else None
                    if actual is None or actual >= len(cells): return None
                    return parse_price(cells[actual].get_text(strip=True))

                ungraded = cell_price(ui)
                grade9   = cell_price(g9i)
                psa10    = cell_price(p10i)

                if ungraded is None and psa10 is None:
                    continue

                # Score by how well the row label matches the requested parallel
                score = 0
                if not label_is_price and parallel_words:
                    label_n = normalize(row_label)
                    score   = sum(1 for w in parallel_words if w not in label_n) * 10

                if score < best_score:
                    best_score = score
                    best = {
                        "title":    page_title if label_is_price else f"{page_title} ({row_label})",
                        "set":      '',
                        "ungraded": ungraded,
                        "grade9":   grade9,
                        "psa10":    psa10,
                        "url":      current_url,
                    }

            if best:
                print(f"  ✓ Grade-table match (score={best_score}): {best['title']}")
            else:
                print("  – No usable rows in grade table")
            return best

        print(f"  ✗ Unrecognised table headers: {hdrs}")
        return None

    except Exception as e:
        print(f"  ✗ Search error: {e}")
        return None


# ── Main fetch ────────────────────────────────────────────────────────────────

def fetch_scp(player, parallel, year=None, card_no=None):
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        page = ctx.new_page()
        result = None
        try:
            # Strategy 1: direct URL (works when year + card_no are provided)
            if year and card_no:
                direct_url = direct_card_url(player, parallel, year, card_no)
                if direct_url:
                    result = scrape_card_page(page, direct_url)

            # Strategy 2: search page fallback
            if not result:
                result = scrape_search_page(page, player, parallel, card_no)

        finally:
            browser.close()

    if result:
        print(
            f"  ✓ Final result: RAW=${result.get('ungraded')}  "
            f"PSA9=${result.get('grade9')}  PSA10=${result.get('psa10')}"
        )
    return result


# ── HTTP handler ──────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        pass

    def send_json(self, code, data):
        body = json.dumps(data).encode()
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", len(body))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_json(204, {})

    def do_POST(self):
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length", 0))
        body   = json.loads(self.rfile.read(length) or b"{}") if length else {}

        if parsed.path in ("/api/update-comps", "/api/update-card"):
            binder     = body.get("binder", "")
            player     = body.get("player", "")
            card_no    = body.get("card_no", "")
            xlsx_row   = body.get("_row")
            comps      = body.get("comps")        # None = not provided, [] = clear
            card_price = body.get("card_price")
            shipping   = body.get("shipping")
            taxes      = body.get("taxes")
            scp_value  = body.get("scp_value")
            tmv        = body.get("tmv")
            new_binder = body.get("new_binder")
            status     = body.get("status")

            try:
                import openpyxl as _xl
                wb = _xl.load_workbook(XLSX)
                ws = wb["Bowman"]
                target = None
                for row in ws.iter_rows(min_row=3):
                    # Primary: exact xlsx row number (never ambiguous)
                    if xlsx_row is not None:
                        if row[0].row == int(xlsx_row):
                            target = row; break
                        continue
                    rv = row[0].value
                    # Fallback 1: binder slot
                    if binder and rv is not None:
                        try:
                            if abs(float(rv) - float(binder)) < 0.001:
                                target = row; break
                        except: pass
                    # Fallback 2: player + card_no
                    if not target and player:
                        rp = str(row[1].value or "").strip().lower()
                        rc = str(row[4].value or "").strip().lower()
                        if rp == player.lower() and rc == card_no.lower():
                            target = row; break

                if not target:
                    self.send_json(404, {"ok": False, "message": "Card not found in xlsx"})
                    wb.close(); return

                # Helper: write by 1-based column number (safe even if row is short)
                def wc(col1, val):
                    ws.cell(row=target[0].row, column=col1, value=val)

                # Update binder slot (col A = 1)
                if new_binder is not None:
                    try:    wc(1, float(new_binder))
                    except: wc(1, new_binder)
                # Write cost fields (cols J=10, K=11, L=12)
                if card_price is not None: wc(10, float(card_price))
                if shipping   is not None: wc(11, float(shipping))
                if taxes      is not None: wc(12, float(taxes))
                # SCP value (col N = 14)
                if scp_value  is not None: wc(14, float(scp_value))
                # TMV (col O = 15) — save explicitly to survive formula-cache stripping
                if tmv is not None: wc(15, float(tmv) if float(tmv) > 0 else None)
                # Status (col V = 22)
                if status is not None:
                    valid = ('sell', 'hold', 'flip')
                    wc(22, status if status in valid else None)
                # Comps (cols Q-U = 17-21) — only write if comps were explicitly provided
                if comps is not None:
                    for i in range(5):
                        val = comps[i] if i < len(comps) else None
                        wc(17 + i, float(val) if val is not None else None)

                wb.save(XLSX)
                wb.close()
                print(f"[UPDATE] {player} {card_no}{' comps→'+str(comps) if comps is not None else ''}")
                if body.get("skip_rebuild"): self.send_json(200, {"ok": True}); return
                # Rebuild HTML so the static DATA snapshot reflects the change
                result = subprocess.run(
                    [sys.executable, str(HERE / 'build_collection.py')],
                    capture_output=True, text=True, cwd=str(HERE)
                )
                if result.returncode == 0:
                    print(f"[REBUILD] OK")
                    # Push to GitHub so GitHub Pages stays in sync
                    git = subprocess.run(
                        ['git', '-C', str(HERE), 'add', 'baseball_collection.html'],
                        capture_output=True, text=True
                    )
                    git = subprocess.run(
                        ['git', '-C', str(HERE), 'commit', '-m',
                         f'Auto: update comps for {player} {card_no}'],
                        capture_output=True, text=True
                    )
                    if 'nothing to commit' in git.stdout + git.stderr:
                        print(f"[GIT] Nothing changed")
                    else:
                        push = subprocess.run(
                            ['git', '-C', str(HERE), 'push'],
                            capture_output=True, text=True
                        )
                        if push.returncode == 0:
                            print(f"[GIT] Pushed to GitHub")
                        else:
                            print(f"[GIT] Push failed: {push.stderr}")
                else:
                    print(f"[REBUILD] FAILED:\n{result.stderr}")
                self.send_json(200, {"ok": True})
            except Exception as e:
                self.send_json(500, {"ok": False, "message": str(e)})
            return

        if parsed.path == "/api/checklist-own":
            sheet = body.get("sheet", "")
            row   = body.get("row")
            col   = body.get("col")
            owned = bool(body.get("owned"))
            if sheet not in ("Bowman 2026", "Bowman 2025") or not row or not col:
                self.send_json(400, {"ok": False, "message": "bad params"}); return
            try:
                import openpyxl as _xl
                wb = _xl.load_workbook(XLSX)
                ws = wb[sheet]
                ws.cell(row=int(row), column=int(col), value=bool(owned))
                wb.save(XLSX)
                wb.close()
                print(f"[CHECKLIST] {sheet} R{row}C{col} → {owned}")
                self.send_json(200, {"ok": True})
            except Exception as e:
                self.send_json(500, {"ok": False, "message": str(e)})
            return

        if parsed.path == "/api/delete-card":
            binder   = body.get("binder", "")
            player   = body.get("player", "").strip()
            card_no  = body.get("card_no", "").strip()
            xlsx_row = body.get("_row")
            try:
                import openpyxl as _xl
                wb = _xl.load_workbook(XLSX)
                ws = wb["Bowman"]
                target_row = None
                for row in ws.iter_rows(min_row=3):
                    if xlsx_row is not None:
                        if row[0].row == int(xlsx_row):
                            target_row = row[0].row; break
                        continue
                    rv = row[0].value
                    if binder and rv is not None:
                        try:
                            if abs(float(rv) - float(binder)) < 0.001:
                                target_row = row[0].row; break
                        except: pass
                    if not target_row and player:
                        rp = str(row[1].value or "").strip().lower()
                        rc = str(row[4].value or "").strip().lower()
                        if rp == player.lower() and rc == card_no.lower():
                            target_row = row[0].row; break
                if not target_row:
                    self.send_json(404, {"ok": False, "message": "Card not found"}); wb.close(); return
                ws.delete_rows(target_row)
                wb.save(XLSX)
                wb.close()
                print(f"[DELETE] Row {target_row}: {player} {card_no}")
                result = subprocess.run(
                    [sys.executable, str(HERE / 'build_collection.py')],
                    capture_output=True, text=True, cwd=str(HERE)
                )
                if result.returncode == 0:
                    git = subprocess.run(['git', '-C', str(HERE), 'add', 'baseball_collection.html'], capture_output=True)
                    git = subprocess.run(['git', '-C', str(HERE), 'commit', '-m', f'Auto: delete {player} {card_no}'], capture_output=True, text=True)
                    if 'nothing to commit' not in git.stdout + git.stderr:
                        subprocess.run(['git', '-C', str(HERE), 'push'], capture_output=True)
                        print(f"[GIT] Pushed")
                self.send_json(200, {"ok": True})
            except Exception as e:
                self.send_json(500, {"ok": False, "message": str(e)})
            return

        if parsed.path == "/api/add-card":
            player     = body.get("player", "").strip()
            year       = body.get("year")
            parallel   = body.get("parallel", "").strip()
            card_no    = body.get("card_no", "").strip()
            type_      = body.get("type", "").strip()
            binder     = body.get("binder", "").strip()
            card_price = body.get("card_price", 0)
            shipping   = body.get("shipping", 0)
            taxes      = body.get("taxes", 0)
            if not player:
                self.send_json(400, {"ok": False, "message": "player is required"}); return
            try:
                import openpyxl as _xl
                wb = _xl.load_workbook(XLSX)
                ws = wb["Bowman"]
                # Find last data row
                last_row = 2
                for row in ws.iter_rows(min_row=3):
                    if any(c.value is not None for c in row):
                        last_row = row[0].row
                new_row = last_row + 1
                # Col A=binder(0), B=player(1), C=year(2), D=parallel(3), E=card_no(4),
                # I=type(8), J=card_price(9), K=shipping(10), L=taxes(11)
                try: binder_val = float(binder) if binder else None
                except: binder_val = binder or None
                ws.cell(new_row, 1, binder_val)
                ws.cell(new_row, 2, player)
                ws.cell(new_row, 3, int(year) if year else None)
                ws.cell(new_row, 4, parallel or None)
                ws.cell(new_row, 5, card_no or None)
                ws.cell(new_row, 9, type_ or None)
                ws.cell(new_row, 10, float(card_price or 0))
                ws.cell(new_row, 11, float(shipping or 0))
                ws.cell(new_row, 12, float(taxes or 0))
                wb.save(XLSX)
                wb.close()
                print(f"[ADD] Row {new_row}: {player} {parallel} {card_no}")
                # Rebuild + push
                result = subprocess.run(
                    [sys.executable, str(HERE / 'build_collection.py')],
                    capture_output=True, text=True, cwd=str(HERE)
                )
                if result.returncode == 0:
                    print(f"[REBUILD] OK")
                    git = subprocess.run(['git', '-C', str(HERE), 'add', 'baseball_collection.html'], capture_output=True)
                    git = subprocess.run(['git', '-C', str(HERE), 'commit', '-m', f'Auto: add {player} {card_no}'], capture_output=True, text=True)
                    if 'nothing to commit' not in git.stdout + git.stderr:
                        push = subprocess.run(['git', '-C', str(HERE), 'push'], capture_output=True, text=True)
                        print(f"[GIT] {'Pushed' if push.returncode==0 else 'Push failed: '+push.stderr}")
                else:
                    print(f"[REBUILD] FAILED:\n{result.stderr}")
                self.send_json(200, {"ok": True})
            except Exception as e:
                self.send_json(500, {"ok": False, "message": str(e)})
            return

        self.send_json(404, {"error": "not found"})

    def do_GET(self):
        parsed = urlparse(self.path)
        qs     = parse_qs(parsed.query)

        if parsed.path == "/health":
            self.send_json(200, {"status": "ok", "port": PORT})
            return

        if parsed.path == "/api/rebuild":
            result = subprocess.run(
                [sys.executable, str(HERE / 'build_collection.py')],
                capture_output=True, text=True, cwd=str(HERE)
            )
            if result.returncode != 0:
                self.send_json(500, {"ok": False, "message": result.stderr})
                return
            git = subprocess.run(['git', '-C', str(HERE), 'add', 'baseball_collection.html'], capture_output=True)
            git = subprocess.run(['git', '-C', str(HERE), 'commit', '-m', 'Auto: bulk SCP update'], capture_output=True, text=True)
            if 'nothing to commit' not in git.stdout + git.stderr:
                subprocess.run(['git', '-C', str(HERE), 'push'], capture_output=True)
            print("[REBUILD] Done via /api/rebuild")
            self.send_json(200, {"ok": True})
            return

        if parsed.path == "/api/scp":
            player   = qs.get("player",   [""])[0].strip()
            parallel = qs.get("parallel", [""])[0].strip()
            year     = qs.get("year",     [""])[0].strip() or None
            card_no  = qs.get("card_no",  [""])[0].strip() or None

            if not player:
                self.send_json(400, {"ok": False, "message": "player is required"})
                return

            print(f"\n[SCP] {player} | {parallel or '—'} | {card_no or '—'} | {year or '—'}")
            t0      = time.time()
            result  = fetch_scp(player, parallel, year=year, card_no=card_no)
            elapsed = round(time.time() - t0, 1)
            print(f"  Done in {elapsed}s")

            if result:
                self.send_json(200, {"ok": True, "data": result, "elapsed": elapsed})
            else:
                self.send_json(200, {
                    "ok":      False,
                    "message": f"No match found on SCP for '{player} {parallel}'",
                })
            return

        self.send_json(404, {"error": "not found"})


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    server = HTTPServer(("localhost", PORT), Handler)
    print(f"\n⚾  Collection Server  →  http://localhost:{PORT}")
    print(f"   Open baseball_collection.html and use the Valuation tab.")
    print(f"   Press Ctrl+C to stop.\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
        server.server_close()
