#!/usr/bin/env python3
"""
build_collection.py — Rebuild the baseball collection dashboard from the xlsx.

Usage:
    python3 build_collection.py               # normal build, uses cached SCP URLs
    python3 build_collection.py --fetch-scp   # fetch missing SCP URLs then build
                                              # (requires collection_server.py running)

Reads:  ~/Downloads/Baseball Collection.xlsx
Writes: collection_data.json  (same folder as this script)
        scp_url_cache.json    (persisted SCP URL cache, never wiped)
        baseball_collection.html  (updates const DATA = {...} in-place)
"""

import json, re, sys, urllib.request, urllib.parse
from pathlib import Path
import openpyxl

XLSX      = Path.home() / 'Downloads' / 'Baseball Collection.xlsx'
HERE      = Path(__file__).parent
HTML      = HERE / 'baseball_collection.html'
JSON_OUT  = HERE / 'collection_data.json'
SCP_CACHE = HERE / 'scp_url_cache.json'
SCP_PORT  = 5055

FETCH_SCP = '--fetch-scp' in sys.argv


def _float(v, default=0.0):
    try:   return float(v) if v is not None else default
    except: return default

def _str(v):
    return str(v).strip() if v is not None else ''

def _binder(v):
    if v is None: return ''
    try:
        f = float(v)
        return f'{f:.1f}'
    except:
        return str(v).strip()

def _date(v):
    if v is None: return ''
    if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
    return str(v)


# ── SCP URL cache ─────────────────────────────────────────────────────────────

def load_scp_cache():
    if SCP_CACHE.exists():
        try: return json.loads(SCP_CACHE.read_text())
        except: pass
    return {}

def save_scp_cache(cache):
    SCP_CACHE.write_text(json.dumps(cache, indent=2))

def scp_cache_key(c):
    return f"{c['player']}|{c['parallel']}|{c['year']}|{c['card_no']}"

def fetch_scp_url(player, parallel, year, card_no):
    """Call the local collection_server to get the direct SCP card URL."""
    params = urllib.parse.urlencode({
        'player':   player,
        'parallel': parallel,
        'year':     year or '',
        'card_no':  card_no or '',
    })
    try:
        with urllib.request.urlopen(
            f'http://localhost:{SCP_PORT}/api/scp?{params}', timeout=40
        ) as r:
            data = json.loads(r.read())
            if data.get('ok') and data.get('data', {}).get('url'):
                return data['data']['url']
    except Exception as e:
        print(f'    ✗ request failed: {e}')
    return None


# ── Sheets ────────────────────────────────────────────────────────────────────

def parse_bowman(ws):
    """Row 1 = merged title, Row 2 = headers, Row 3+ = card data."""
    cards = []
    for xlsx_row, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        player = _str(row[1])
        if not player:
            continue
        comps = [round(_float(row[i]), 2) for i in range(16, 21) if row[i] is not None and _float(row[i]) > 0]
        card_price = round(_float(row[9]),  2)
        shipping   = round(_float(row[10]), 2)
        taxes      = round(_float(row[11]), 2)
        # row[12] is a formula (=card+ship+tax); cache may be stripped by openpyxl —
        # fall back to computing it from components
        cost_raw   = _float(row[12])
        cost       = round(cost_raw if cost_raw else card_price + shipping + taxes, 2)
        scp_value  = round(_float(row[13]), 2)
        # row[14] may be a formula (e.g. AVERAGE of comps); fall back to avg of comps
        tmv_raw    = _float(row[14])
        if tmv_raw:
            tmv = round(tmv_raw, 2)
        elif comps:
            tmv = round(sum(comps) / len(comps), 2)
        else:
            tmv = 0.0
        # row[15] is also a formula; recompute if cache is missing
        pl_raw     = row[15]
        if pl_raw is None or isinstance(pl_raw, str):
            pl = round(((tmv - cost) / cost * 100) if cost else 0.0, 2)
        else:
            pl = round(_float(pl_raw), 2)
        # Status: column V (index 21) — hold / flip / watching / ''
        status_raw = row[21] if len(row) > 21 else None
        status = _str(status_raw).lower() if status_raw else ''
        if status not in ('sell', 'hold', 'flip'):
            status = ''
        cards.append({
            'binder':     _binder(row[0]),
            'player':     player,
            'year':       int(row[2]) if row[2] else None,
            'parallel':   _str(row[3]),
            'card_no':    _str(row[4]),
            'type':       _str(row[8]),
            'card_price': card_price,
            'shipping':   shipping,
            'taxes':      taxes,
            'cost':       cost,
            'scp_value':  scp_value,
            'tmv':        tmv,
            'pl':         pl,
            'pl_dollars': round(tmv - cost, 2),
            'comps':      comps,
            'status':     status,
            '_row':       xlsx_row,   # exact xlsx row number for reliable matching
            'scp_url':    None,  # filled in below from cache
        })
    return cards


def parse_fav_players(ws):
    """Row 1 = headers, Row 2+ = data."""
    favs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        favs.append({
            'player':      _str(row[0]),
            'team':        _str(row[1]),
            'rank':        _float(row[2]) if row[2] else None,
            'level':       _str(row[3]),
            'favorite':    bool(row[4]),
            'rookie_year': _str(row[5]),
        })
    return favs


def parse_watchlist(ws):
    """Row 1 = headers, Row 2+ = data."""
    wl = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        wl.append({
            'player':     _str(row[0]),
            'price_tier': _str(row[1]),
        })
    return wl


def parse_transactions(ws):
    """
    Row 1: section labels  (Ripping Wax | Slinging Singles)
    Row 2: column headers  (Product, Price, Date | Player, Product, Price, Date)
    Row 3: totals row      (Total, <sum>, ... | Total, None, <sum>, ...)
    Row 4+: individual transactions
    """
    wax, singles = [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if _str(row[0]).lower() == 'total':
            continue  # skip totals row; we sum dynamically below
        if row[0] and row[1] is not None:
            wax.append({
                'product': _str(row[0]),
                'price':   round(_float(row[1]), 2),
                'date':    _date(row[2]),
            })
        if row[3] and row[5] is not None:
            singles.append({
                'player':  _str(row[3]),
                'product': _str(row[4]),
                'price':   round(_float(row[5]), 2),
                'date':    _date(row[6]),
            })
    return wax, singles


def parse_rookies(ws):
    """Row 1 = 'Favorite Rookie' label, Row 2 = headers, Row 3+ = data."""
    rookies = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        rookies.append({
            'player':     _str(row[0]),
            'team':       _str(row[1]),
            'year':       int(row[2]) if row[2] else None,
            'card_title': _str(row[3]),
            'set':        _str(row[4]),
            'card_no':    _str(row[5]),
            'value':      round(_float(row[6]), 2),
            'collected':  bool(row[7]),
            'wishlist':   bool(row[8]),
        })
    return rookies


# Friendly display names per card-number prefix. Edit freely — unknown
# prefixes fall back to the raw prefix string.
SET_NAMES = {
    'BASE': 'Base Set',
    'BP':   'Paper Prospects',
    'BCP':  'Chrome Prospects',
    'BTP':  'Scouts Top 100',
    'BST':  'Bowman Sterling',
    'ES':   'Electric Sluggers',
    'UR':   'Under the Radar',
    'PC':   'Power Chords',
    'ROY':  'ROY Favorites',
    'GL':   'Greatness Loading',
    'RR':   'Rockstar Rookies',
    'VIP':  'Very Important Prospects',
}

def parse_checklist(ws):
    """
    Multi-column layout with repeating groups of (#, Player, Team, [Year,] Own, spacer).
    Each column group is one insert set (Base, BP, BCP, ES, ...).
    Returns per-set breakdown with exact xlsx cell coordinates for live editing.
    """
    rows = list(ws.iter_rows())
    if not rows:
        return {'total': 0, 'owned': 0, 'sets': []}

    headers = [c.value for c in rows[0]]
    groups  = []  # (no_col, player_col, team_col, own_col)  all 0-based

    i = 0
    while i < len(headers):
        if headers[i] == '#' and i + 1 < len(headers) and headers[i + 1] == 'Player':
            team_col = i + 2 if i + 2 < len(headers) and headers[i + 2] == 'Team' else None
            for j in range(i + 2, min(i + 7, len(headers))):
                if headers[j] == 'Own':
                    groups.append((i, i + 1, team_col, j))
                    i = j + 2
                    break
            else:
                i += 1
        else:
            i += 1

    sets = []
    for gi, (no_col, player_col, team_col, own_col) in enumerate(groups):
        cards  = []
        prefix = None
        for r in rows[1:]:
            if player_col >= len(r) or not r[player_col].value:
                continue
            card_no  = _str(r[no_col].value)
            player   = _str(r[player_col].value)
            team     = _str(r[team_col].value) if team_col is not None and team_col < len(r) else ''
            own_cell = r[own_col] if own_col < len(r) else None
            is_owned = bool(own_cell.value) if own_cell is not None and own_cell.value is not None else False
            if prefix is None and card_no:
                m = re.match(r'^([A-Za-z]+)-', card_no)
                prefix = m.group(1).upper() if m else 'BASE'
            cards.append({
                'no':     card_no,
                'player': player,
                'team':   team,
                'owned':  is_owned,
                'row':    r[no_col].row,   # xlsx row (1-based)
                'col':    own_col + 1,     # xlsx Own column (1-based)
            })
        if not cards:
            continue
        if prefix is None:
            prefix = f'SET{gi+1}'
        sets.append({
            'key':    f'{prefix}_{gi}',
            'prefix': prefix,
            'name':   SET_NAMES.get(prefix, prefix),
            'total':  len(cards),
            'owned':  sum(1 for c in cards if c['owned']),
            'cards':  cards,
        })

    total = sum(s['total'] for s in sets)
    owned = sum(s['owned'] for s in sets)
    return {'total': total, 'owned': owned, 'sets': sets}


# ── Main ──────────────────────────────────────────────────────────────────────

def build():
    if not XLSX.exists():
        print(f'✗  Not found: {XLSX}')
        sys.exit(1)

    print(f'Reading {XLSX.name} ...')
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    cards        = parse_bowman(wb['Bowman'])
    fav_players  = parse_fav_players(wb['Favorite Players']) if 'Favorite Players' in wb.sheetnames else []
    watchlist    = parse_watchlist(wb['Watchlist'])
    wax, singles = parse_transactions(wb['Transactions'])
    rookies      = parse_rookies(wb['Rookies'])
    cl_2026      = parse_checklist(wb['Bowman 2026'])
    cl_2025      = parse_checklist(wb['Bowman 2025'])
    wb.close()

    # ── SCP URL cache ──────────────────────────────────────────────────────────
    scp_cache = load_scp_cache()

    if FETCH_SCP:
        # Check server is reachable first
        try:
            urllib.request.urlopen(f'http://localhost:{SCP_PORT}/health', timeout=3)
        except Exception:
            print(f'✗  collection_server.py not running on :{SCP_PORT}')
            print(f'   Start it first: python3 collection_server.py')
            sys.exit(1)

        missing = [c for c in cards if scp_cache_key(c) not in scp_cache]
        print(f'\nFetching SCP URLs for {len(missing)} uncached cards...')
        for i, c in enumerate(missing, 1):
            key = scp_cache_key(c)
            label = f"{c['player']} {c['parallel']}"
            print(f'  [{i}/{len(missing)}] {label}')
            url = fetch_scp_url(c['player'], c['parallel'], c['year'], c['card_no'])
            scp_cache[key] = url  # store None for misses so we don't retry
            if url:
                print(f'    ✓ {url}')
            else:
                print(f'    – not found')
        save_scp_cache(scp_cache)
        cached_count = sum(1 for v in scp_cache.values() if v)
        print(f'✓  scp_url_cache.json  ({cached_count} URLs cached)\n')

    # Apply cached URLs to cards
    for c in cards:
        c['scp_url'] = scp_cache.get(scp_cache_key(c))

    # ── Summaries ──────────────────────────────────────────────────────────────
    wax_spent     = round(sum(t['price'] for t in wax), 2)
    singles_spent = round(sum(t['price'] for t in singles), 2)
    total_cost    = round(sum(c['cost']       for c in cards), 2)
    total_tmv     = round(sum(c['tmv']        for c in cards), 2)
    total_pl      = round(sum(c['pl_dollars'] for c in cards), 2)
    total_pl_pct  = round(((total_tmv - total_cost) / total_cost * 100) if total_cost else 0.0, 1)

    data = {
        'summary': {
            'total_cards':   len(cards),
            'total_cost':    total_cost,
            'total_tmv':     total_tmv,
            'total_pl':      total_pl,
            'total_pl_pct':  total_pl_pct,
            'wax_spent':     wax_spent,
            'singles_spent': singles_spent,
        },
        'cards':                cards,
        'fav_players':          fav_players,
        'watchlist':            watchlist,
        'wax_transactions':     wax,
        'singles_transactions': singles,
        'rookies':              rookies,
        'checklist': {
            'bowman_2026': cl_2026,
            'bowman_2025': cl_2025,
        },
    }

    # Write JSON sidecar
    JSON_OUT.write_text(json.dumps(data, indent=2))
    print(f'✓  {JSON_OUT.name}  ({len(cards)} cards)')

    # Patch the HTML — replace const DATA = {...};
    if not HTML.exists():
        print(f'✗  HTML not found: {HTML}')
        sys.exit(1)

    html = HTML.read_text()
    new_block = f'const DATA = {json.dumps(data, separators=(",", ":"))};'
    patched, n = re.subn(r'const DATA = \{.*?\};', lambda _: new_block, html, flags=re.DOTALL)
    if n == 0:
        print('✗  Could not locate "const DATA = {...}" in HTML — file not updated')
        sys.exit(1)

    HTML.write_text(patched)
    print(f'✓  {HTML.name}  (DATA block replaced)')

    urls_in_data = sum(1 for c in cards if c.get('scp_url'))
    print(f'\n{"─"*46}')
    print(f'  Cards        {len(cards):>6}')
    print(f'  SCP URLs     {urls_in_data:>6} / {len(cards)}')
    print(f'  Total Cost   ${total_cost:>9.2f}')
    print(f'  Total TMV    ${total_tmv:>9.2f}')
    print(f'  P/L          ${total_pl:>+9.2f}')
    print(f'  Wax spent    ${wax_spent:>9.2f}')
    print(f'  Singles      ${singles_spent:>9.2f}')
    print(f'  CL 2026      {cl_2026["owned"]:>3}/{cl_2026["total"]}')
    print(f'  CL 2025      {cl_2025["owned"]:>3}/{cl_2025["total"]}')
    print(f'{"─"*46}')
    if FETCH_SCP:
        print('  Done. Run again without --fetch-scp for faster builds.')
    else:
        print('  Done. Reload baseball_collection.html in your browser.')
        if urls_in_data < len(cards):
            print(f'  Tip: run with --fetch-scp to fill in {len(cards)-urls_in_data} missing SCP links.')


if __name__ == '__main__':
    build()
